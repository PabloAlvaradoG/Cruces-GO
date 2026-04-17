import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import io
import time
import bisect
from datetime import datetime

# ══════════════════════════════════════════════════════════════════════════════
# PAGE CONFIG
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Motor de Cruces de Inventario",
    page_icon="📦",
    layout="wide"
)

# ══════════════════════════════════════════════════════════════════════════════
# ESTILOS
# ══════════════════════════════════════════════════════════════════════════════
AZUL_OSC  = "1F3864"
AZUL_MED  = "2E75B6"
AZUL_CLR  = "DDEBF7"
ROJO_HDR  = "C00000"
ROJO_CLR  = "FCE4D6"
ROJO_ALT  = "F8CBAD"
VERDE_HDR = "375623"
VERDE_CLR = "E2EFDA"
VERDE_ALT = "C6E0B4"
AMARILLO  = "FFC000"
AMAR_CLR  = "FFF2CC"
GRIS_CLR  = "F2F2F2"
GRIS_MED  = "D9D9D9"
BLANCO    = "FFFFFF"
NEGRO     = "000000"

thin = Side(style="thin", color="BFBFBF")
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_c):    return PatternFill("solid", fgColor=hex_c)
def font(bold=False, color=NEGRO, sz=10): return Font(name="Arial", bold=bold, size=sz, color=color)
def alig(h="left", wrap=True): return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def set_col_width(ws, col, width): ws.column_dimensions[get_column_letter(col)].width = width

def write_header(ws, row, col, value, bg, fg=BLANCO, sz=10, bold=True, h="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, color=fg, sz=sz)
    c.fill = fill(bg); c.border = brd; c.alignment = alig(h)
    return c

def write_cell(ws, row, col, value, bg=BLANCO, fg=NEGRO, bold=False, sz=9, h="left", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, color=fg, sz=sz)
    c.fill = fill(bg); c.border = brd; c.alignment = alig(h)
    if num_fmt: c.number_format = num_fmt
    return c

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACION POR DEFECTO
# ══════════════════════════════════════════════════════════════════════════════
MAPA_COLUMNAS = {
    "Material":       ["Material", "material", "Codigo", "Código"],
    "Descripcion":    ["Descripcion", "Descripción", "descripcion"],
    "Categoria":      ["Categoria", "Categoría", "categoria", "Familia"],
    "Diferencia":     ["Diferencia", "diferencia", "Cant. Dif"],
    "Costo_Unitario": ["Costo Unitario", "Costo_Unitario", "costo unitario", "Precio U"],
    "Costo_Total":    ["Costo Total", "Costo_Total", "costo total", "Subtotal"],
}

STOPWORDS = {
    'DE','PARA','CON','EL','LA','UN','EN','DEL','LOS','LAS',
    'PZ','MZ','Y','A','AL','KG','MTS','PZA','X'
}

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES AUXILIARES
# ══════════════════════════════════════════════════════════════════════════════
def resolver_columnas(df_cols):
    resultado = {}
    for nombre_interno, variantes in MAPA_COLUMNAS.items():
        encontrado = None
        for v in variantes:
            if v in df_cols:
                encontrado = v
                break
        resultado[nombre_interno] = encontrado
    return resultado

def obtener_palabras(texto):
    if pd.isna(texto) or str(texto).strip() == "":
        return set()
    txt = str(texto).upper()
    txt = re.sub(r'[/\-_]', ' ', txt)
    txt = re.sub(r'[^A-Z0-9\s]', '', txt)
    resultado = set()
    for p in txt.split():
        if len(p) <= 2: continue
        if p in STOPWORDS: continue
        digitos = sum(1 for c in p if c.isdigit())
        if digitos / len(p) > 0.5: continue
        resultado.add(p)
    return resultado

def jaccard(set1, set2):
    if not set1 or not set2: return 0.0
    return len(set1 & set2) / len(set1 | set2)

def nivel_confianza(sim, alta_min, media_min):
    if sim >= alta_min:  return ("ALTA",  "🟢", VERDE_CLR)
    if sim >= media_min: return ("MEDIA", "🟡", AMAR_CLR)
    return ("BAJA", "🔴", ROJO_CLR)

# ══════════════════════════════════════════════════════════════════════════════
# CARGA Y PROCESAMIENTO
# ══════════════════════════════════════════════════════════════════════════════
def cargar_configuracion(file_bytes):
    cfg_defaults = {
        "TOLERANCIA_COSTO": 0.05,
        "UMBRAL_JACCARD":   0.30,
        "ALTA_MIN":         0.70,
        "MEDIA_MIN":        0.45,
        "FILTRAR_CATEGORIA": True,
        "GENERAR_DIAG":     False,
        "nombre_tienda":    "",
        "centro_sap":       "",
        "fecha_inventario": "",
    }
    try:
        df_cfg = pd.read_excel(file_bytes, sheet_name="Configuracion", header=None)
        cfg = {}
        for _, row in df_cfg.iterrows():
            if pd.notna(row.iloc[0]) and pd.notna(row.iloc[1]):
                cfg[str(row.iloc[0]).strip()] = row.iloc[1]

        if "Tolerancia Costo %"       in cfg: cfg_defaults["TOLERANCIA_COSTO"]  = float(cfg["Tolerancia Costo %"]) / 100
        if "Umbral Similitud %"       in cfg: cfg_defaults["UMBRAL_JACCARD"]    = float(cfg["Umbral Similitud %"]) / 100
        if "Umbral Confianza Alta %"  in cfg: cfg_defaults["ALTA_MIN"]          = float(cfg["Umbral Confianza Alta %"]) / 100
        if "Umbral Confianza Media %" in cfg: cfg_defaults["MEDIA_MIN"]         = float(cfg["Umbral Confianza Media %"]) / 100
        if "Filtrar por Categoria"    in cfg: cfg_defaults["FILTRAR_CATEGORIA"] = str(cfg["Filtrar por Categoria"]).strip().upper() == "SI"
        if "Generar Diagnostico"      in cfg: cfg_defaults["GENERAR_DIAG"]      = str(cfg["Generar Diagnostico"]).strip().upper() == "SI"
        if "Nombre Tienda"            in cfg: cfg_defaults["nombre_tienda"]     = str(cfg["Nombre Tienda"]).strip()
        if "Centro SAP"               in cfg: cfg_defaults["centro_sap"]        = str(cfg["Centro SAP"]).strip()
        if "Fecha Inventario"         in cfg: cfg_defaults["fecha_inventario"]  = str(cfg["Fecha Inventario"]).strip()
    except Exception:
        pass
    return cfg_defaults

def cargar_datos(file_bytes):
    df = pd.read_excel(file_bytes, sheet_name="Diferencias", header=1, dtype={"Material": str})
    mapa = resolver_columnas(df.columns.tolist())

    obligatorias = ["Material", "Descripcion", "Diferencia", "Costo_Unitario", "Costo_Total"]
    faltantes_col = [k for k in obligatorias if mapa[k] is None]
    if faltantes_col:
        raise ValueError(f"Columnas no encontradas: {faltantes_col}\nColumnas en archivo: {df.columns.tolist()}")

    rename_map = {v: k for k, v in mapa.items() if v is not None}
    df = df.rename(columns=rename_map)
    if "Categoria" not in df.columns: df["Categoria"] = ""
    df = df[df["Diferencia"] != 0].copy()
    df["Diferencia"]     = pd.to_numeric(df["Diferencia"],     errors="coerce").fillna(0)
    df["Costo_Unitario"] = pd.to_numeric(df["Costo_Unitario"], errors="coerce").fillna(0)
    df["Costo_Total"]    = pd.to_numeric(df["Costo_Total"],    errors="coerce").fillna(0)
    df["Categoria"]      = df["Categoria"].fillna("").astype(str).str.strip()
    return df

def ejecutar_cruces(df, cfg, progress_bar, status_text):
    TOLERANCIA_COSTO  = cfg["TOLERANCIA_COSTO"]
    UMBRAL_JACCARD    = cfg["UMBRAL_JACCARD"]
    ALTA_MIN          = cfg["ALTA_MIN"]
    MEDIA_MIN         = cfg["MEDIA_MIN"]
    FILTRAR_CATEGORIA = cfg["FILTRAR_CATEGORIA"]

    faltantes = df[df["Diferencia"] < 0].copy().reset_index(drop=True)
    sobrantes = df[df["Diferencia"] > 0].copy().reset_index(drop=True)

    faltantes["Saldo"]    = faltantes["Diferencia"].abs()
    sobrantes["Saldo"]    = sobrantes["Diferencia"].abs()
    faltantes["Palabras"] = faltantes["Descripcion"].apply(obtener_palabras)
    sobrantes["Palabras"] = sobrantes["Descripcion"].apply(obtener_palabras)

    sob_costos       = sobrantes["Costo_Unitario"].values
    sob_sorted       = sob_costos.argsort()
    sob_costos_sorted = sob_costos[sob_sorted]

    acum_f = {i: 0.0 for i in range(len(faltantes))}
    acum_s = {i: 0.0 for i in range(len(sobrantes))}

    cruces   = []
    log_diag = []
    contador = 0
    total_f  = len(faltantes)

    # ── Pasada 1: con validación de costo ────────────────────────────────────
    status_text.text("⚙️ Pasada 1 — cruzando con validación de costo...")
    for idx_f in range(total_f):
        if faltantes.at[idx_f, "Saldo"] <= 0: continue

        pct = (idx_f + 1) / total_f * 0.5
        progress_bar.progress(pct, text=f"Pasada 1: {idx_f+1}/{total_f} faltantes | Cruces: {contador}")

        costo_f    = faltantes.at[idx_f, "Costo_Unitario"]
        lim_inf    = costo_f * (1 - TOLERANCIA_COSTO)
        lim_sup    = costo_f * (1 + TOLERANCIA_COSTO)
        palabras_f = faltantes.at[idx_f, "Palabras"]
        cat_f      = faltantes.at[idx_f, "Categoria"]
        mat_f      = faltantes.at[idx_f, "Material"]
        desc_f     = faltantes.at[idx_f, "Descripcion"]

        i_ini = bisect.bisect_left(sob_costos_sorted,  lim_inf)
        i_fin = bisect.bisect_right(sob_costos_sorted, lim_sup)
        candidatos_idx = sob_sorted[i_ini:i_fin]

        candidatos = []
        for idx_s in candidatos_idx:
            if sobrantes.at[idx_s, "Saldo"] <= 0: continue
            costo_s = sobrantes.at[idx_s, "Costo_Unitario"]
            if not (lim_inf <= costo_s <= lim_sup): continue
            if FILTRAR_CATEGORIA:
                c1 = cat_f.strip(); c2 = sobrantes.at[idx_s, "Categoria"].strip()
                if c1 != "" and c2 != "" and c1.upper() != c2.upper(): continue
            sim = jaccard(palabras_f, sobrantes.at[idx_s, "Palabras"])
            if sim < UMBRAL_JACCARD: continue
            candidatos.append((idx_s, sim))

        candidatos.sort(key=lambda x: x[1], reverse=True)

        for idx_s, sim in candidatos:
            if faltantes.at[idx_f, "Saldo"] <= 0: break
            cant      = min(faltantes.at[idx_f, "Saldo"], sobrantes.at[idx_s, "Saldo"])
            costo_u_s = sobrantes.at[idx_s, "Costo_Unitario"]
            nivel, emoji, _ = nivel_confianza(sim, ALTA_MIN, MEDIA_MIN)
            contador += 1

            dif_cruce_f = round(-cant, 2)
            dif_cruce_s = round( cant, 2)
            ct_f        = round(-cant * costo_f,   2)
            ct_s        = round( cant * costo_u_s, 2)
            dif_cant    = round(dif_cruce_f + dif_cruce_s, 2)
            dif_costo   = round(ct_f + ct_s, 2)
            acum_f[idx_f] += cant
            acum_s[idx_s] += cant

            palabras_s = sobrantes.at[idx_s, "Palabras"]
            cruces.append({
                "No_Cruce": contador, "Confianza": emoji + " " + nivel,
                "Similitud": round(sim * 100, 1),
                "Comunes": ", ".join(sorted(palabras_f & palabras_s)),
                "Unicas":  ", ".join(sorted(palabras_f | palabras_s)),
                "Mat_F": mat_f, "Desc_F": desc_f, "Cat_F": cat_f,
                "Dif_F": dif_cruce_f, "CostoU_F": costo_f, "CostoT_F": ct_f,
                "Cantidad_Cruce": cant,
                "Mat_S": sobrantes.at[idx_s, "Material"],
                "Desc_S": sobrantes.at[idx_s, "Descripcion"],
                "Cat_S":  sobrantes.at[idx_s, "Categoria"],
                "Dif_S": dif_cruce_s, "CostoU_S": costo_u_s, "CostoT_S": ct_s,
                "Dif_Cantidad": dif_cant, "Dif_Costo": dif_costo,
            })
            faltantes.at[idx_f, "Saldo"] -= cant
            sobrantes.at[idx_s, "Saldo"] -= cant

    # ── Pasada 2: sin validación de costo ────────────────────────────────────
    status_text.text("⚙️ Pasada 2 — cruzando sin validación de costo...")
    cruces_sc  = []
    contador_sc = 0
    total_f2    = sum(1 for i in range(len(faltantes)) if faltantes.at[i, "Saldo"] > 0)
    proc_f2     = 0

    for idx_f in range(len(faltantes)):
        if faltantes.at[idx_f, "Saldo"] <= 0: continue
        proc_f2 += 1
        pct2 = 0.5 + (proc_f2 / max(total_f2, 1)) * 0.4
        progress_bar.progress(pct2, text=f"Pasada 2: {proc_f2}/{total_f2} | Cruces sin costo: {contador_sc}")

        costo_f    = faltantes.at[idx_f, "Costo_Unitario"]
        palabras_f = faltantes.at[idx_f, "Palabras"]
        cat_f      = faltantes.at[idx_f, "Categoria"]
        mat_f      = faltantes.at[idx_f, "Material"]
        desc_f     = faltantes.at[idx_f, "Descripcion"]

        candidatos_sc = []
        for idx_s in range(len(sobrantes)):
            if sobrantes.at[idx_s, "Saldo"] <= 0: continue
            if FILTRAR_CATEGORIA:
                c1 = cat_f.strip(); c2 = sobrantes.at[idx_s, "Categoria"].strip()
                if c1 != "" and c2 != "" and c1.upper() != c2.upper(): continue
            sim = jaccard(palabras_f, sobrantes.at[idx_s, "Palabras"])
            if sim < UMBRAL_JACCARD: continue
            candidatos_sc.append((idx_s, sim))

        candidatos_sc.sort(key=lambda x: x[1], reverse=True)

        for idx_s, sim in candidatos_sc:
            if faltantes.at[idx_f, "Saldo"] <= 0: break
            cant       = min(faltantes.at[idx_f, "Saldo"], sobrantes.at[idx_s, "Saldo"])
            costo_u_s  = sobrantes.at[idx_s, "Costo_Unitario"]
            nivel, emoji, _ = nivel_confianza(sim, ALTA_MIN, MEDIA_MIN)
            contador_sc += 1

            dif_cruce_f  = round(-cant, 2)
            dif_cruce_s  = round( cant, 2)
            ct_f         = round(-cant * costo_f,   2)
            ct_s         = round( cant * costo_u_s, 2)
            dif_cant     = round(dif_cruce_f + dif_cruce_s, 2)
            dif_costo    = round(ct_f + ct_s, 2)
            dif_costo_pct = round(abs(costo_f - costo_u_s) / (costo_f or 1) * 100, 1)
            acum_f[idx_f] += cant
            acum_s[idx_s] += cant

            palabras_s = sobrantes.at[idx_s, "Palabras"]
            cruces_sc.append({
                "No_Cruce": contador_sc, "Confianza": emoji + " " + nivel,
                "Similitud": round(sim * 100, 1),
                "Comunes": ", ".join(sorted(palabras_f & palabras_s)),
                "Unicas":  ", ".join(sorted(palabras_f | palabras_s)),
                "Mat_F": mat_f, "Desc_F": desc_f, "Cat_F": cat_f,
                "Dif_F": dif_cruce_f, "CostoU_F": costo_f, "CostoT_F": ct_f,
                "Cantidad_Cruce": cant,
                "Mat_S": sobrantes.at[idx_s, "Material"],
                "Desc_S": sobrantes.at[idx_s, "Descripcion"],
                "Cat_S":  sobrantes.at[idx_s, "Categoria"],
                "Dif_S": dif_cruce_s, "CostoU_S": costo_u_s, "CostoT_S": ct_s,
                "Dif_Cantidad": dif_cant, "Dif_Costo": dif_costo,
                "Dif_Costo_Pct": dif_costo_pct,
            })
            faltantes.at[idx_f, "Saldo"] -= cant
            sobrantes.at[idx_s, "Saldo"] -= cant

    # ── Sin cruce final ───────────────────────────────────────────────────────
    no_cruce_final = []
    for idx_f in range(len(faltantes)):
        saldo_rest = faltantes.at[idx_f, "Saldo"]
        if saldo_rest > 0:
            row = faltantes.iloc[idx_f]
            no_cruce_final.append({
                "Material": row["Material"], "Descripcion": row["Descripcion"],
                "Categoria": row["Categoria"], "Diferencia": row["Diferencia"],
                "Costo_Unitario": row["Costo_Unitario"], "Costo_Total": row["Costo_Total"],
                "Tipo": "Faltante", "Cantidad_Disponible": round(-saldo_rest, 2),
            })
    for idx_s in range(len(sobrantes)):
        saldo_rest = sobrantes.at[idx_s, "Saldo"]
        if saldo_rest > 0:
            row = sobrantes.iloc[idx_s]
            no_cruce_final.append({
                "Material": row["Material"], "Descripcion": row["Descripcion"],
                "Categoria": row["Categoria"], "Diferencia": row["Diferencia"],
                "Costo_Unitario": row["Costo_Unitario"], "Costo_Total": row["Costo_Total"],
                "Tipo": "Sobrante", "Cantidad_Disponible": round(saldo_rest, 2),
            })

    resumen_f = {faltantes.at[i, "Material"]: round(acum_f[i], 2) for i in range(len(faltantes))}
    resumen_s = {sobrantes.at[i, "Material"]: round(acum_s[i], 2) for i in range(len(sobrantes))}

    progress_bar.progress(0.9, text="📊 Generando Excel de resultados...")
    return pd.DataFrame(cruces), pd.DataFrame(cruces_sc), pd.DataFrame(no_cruce_final), resumen_f, resumen_s

# ══════════════════════════════════════════════════════════════════════════════
# GENERACION EXCEL RESULTADO
# ══════════════════════════════════════════════════════════════════════════════
def generar_excel(df_input, df_cruces, df_cruces_sc, df_no_cruce, resumen_f, resumen_s, tiempo_seg, cfg):
    wb = openpyxl.Workbook()
    ALTA_MIN  = cfg["ALTA_MIN"]
    MEDIA_MIN = cfg["MEDIA_MIN"]
    info = {k: cfg.get(k, "") for k in ["nombre_tienda","centro_sap","fecha_inventario"]}

    # ── Hoja 1: Diferencias ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Diferencias"
    ws1.sheet_view.showGridLines = False
    ws1.row_dimensions[1].height = 30
    ws1.row_dimensions[2].height = 20
    ws1.merge_cells("A1:I1")
    write_header(ws1, 1, 1, "CONTRALORIA GOH - DIFERENCIAS DE INVENTARIO", AZUL_OSC, sz=12)

    hdrs1 = ["Material","Descripcion","Categoria","Diferencia","Costo Unitario",
             "Costo Total","Cantidad Cruzada","Cantidad Neta","Costo Total S/Cruces"]
    for i, h in enumerate(hdrs1, 1):
        write_header(ws1, 2, i, h, AZUL_MED)

    for i, w in enumerate([18, 45, 18, 14, 16, 16, 18, 16, 20], 1):
        set_col_width(ws1, i, w)

    for r_idx, (_, row) in enumerate(df_input.iterrows(), start=3):
        bg = ROJO_CLR if row["Diferencia"] < 0 else VERDE_CLR
        mat = row["Material"]; dif = row["Diferencia"]; cu = row["Costo_Unitario"]
        cant_cruzada = round(resumen_f.get(mat, 0.0), 2) if dif < 0 else round(-resumen_s.get(mat, 0.0), 2)
        cant_neta    = round(dif + cant_cruzada, 2)
        costo_snc    = round(cant_neta * cu, 2)
        vals = [mat, row["Descripcion"], row["Categoria"], dif, cu,
                row["Costo_Total"], cant_cruzada, cant_neta, costo_snc]
        for c_idx, val in enumerate(vals, 1):
            c = write_cell(ws1, r_idx, c_idx, val, bg=bg)
            if c_idx in (4,5,6,7,8,9): c.number_format = '#,##0.00'
    ws1.freeze_panes = "A3"

    # ── Hoja 2: Cruces_Sugeridos ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Cruces_Sugeridos")
    ws2.sheet_view.showGridLines = False
    ws2.row_dimensions[1].height = 36
    hdrs2 = [
        ("No. Cruce",AZUL_OSC,BLANCO),("Confianza",AZUL_OSC,BLANCO),("% Similitud",AZUL_OSC,BLANCO),
        ("Palabras Comunes",AZUL_CLR,NEGRO),("Total Palabras Unicas",AZUL_CLR,NEGRO),
        ("Material",ROJO_HDR,BLANCO),("Descripcion",ROJO_HDR,BLANCO),("Categoria",ROJO_HDR,BLANCO),
        ("Diferencia",ROJO_HDR,BLANCO),("Costo Unitario",ROJO_HDR,BLANCO),("Costo Total",ROJO_HDR,BLANCO),
        ("Cantidad Cruce",AZUL_OSC,BLANCO),
        ("Material",VERDE_HDR,BLANCO),("Descripcion",VERDE_HDR,BLANCO),("Categoria",VERDE_HDR,BLANCO),
        ("Diferencia",VERDE_HDR,BLANCO),("Costo Unitario",VERDE_HDR,BLANCO),("Costo Total",VERDE_HDR,BLANCO),
        ("Dif. Cantidad",AZUL_MED,BLANCO),("Dif. Costo",AZUL_MED,BLANCO),
    ]
    for i, (h,bg,fg) in enumerate(hdrs2, 1):
        write_header(ws2, 1, i, h, bg, fg=fg)
    for i, w in enumerate([10,14,12,28,32,18,38,16,14,14,14,12,18,38,16,14,14,14,14,14], 1):
        set_col_width(ws2, i, w)
    ws2.freeze_panes = "A2"

    if not df_cruces.empty:
        for r_idx, row in df_cruces.iterrows():
            er  = r_idx + 2; alt = r_idx % 2 == 0
            sim_val = row["Similitud"]
            _, _, color_sem = nivel_confianza(sim_val / 100, ALTA_MIN, MEDIA_MIN)
            bg_f = ROJO_ALT if alt else ROJO_CLR
            bg_s = VERDE_ALT if alt else VERDE_CLR
            dif_costo = row["Dif_Costo"]
            bg_dc = VERDE_CLR if abs(dif_costo) < 0.01 else AMAR_CLR
            vals_row = [
                (row["No_Cruce"],AMAR_CLR,"center",None),(row["Confianza"],color_sem,"center",None),
                (str(sim_val)+"%",color_sem,"center",None),(row["Comunes"],AZUL_CLR,"left",None),
                (row["Unicas"],AZUL_CLR,"left",None),(row["Mat_F"],bg_f,"left",None),
                (row["Desc_F"],bg_f,"left",None),(row["Cat_F"],bg_f,"left",None),
                (row["Dif_F"],bg_f,"center","#,##0.00"),(row["CostoU_F"],bg_f,"center","#,##0.00"),
                (row["CostoT_F"],bg_f,"center","#,##0.00"),(row["Cantidad_Cruce"],AZUL_OSC,"center","#,##0.00"),
                (row["Mat_S"],bg_s,"left",None),(row["Desc_S"],bg_s,"left",None),
                (row["Cat_S"],bg_s,"left",None),(row["Dif_S"],bg_s,"center","#,##0.00"),
                (row["CostoU_S"],bg_s,"center","#,##0.00"),(row["CostoT_S"],bg_s,"center","#,##0.00"),
                (row["Dif_Cantidad"],VERDE_CLR,"center","#,##0.00"),(dif_costo,bg_dc,"center","#,##0.00"),
            ]
            for c_idx, (val,bg,ha,nf) in enumerate(vals_row, 1):
                c = write_cell(ws2, er, c_idx, val, bg=bg, h=ha)
                if nf: c.number_format = nf

    # ── Hoja 3: Cruces_Sin_Costo ──────────────────────────────────────────────
    ws_sc = wb.create_sheet("Cruces_Sin_Costo")
    ws_sc.sheet_view.showGridLines = False
    ws_sc.row_dimensions[1].height = 36
    hdrs_sc = [
        ("No. Cruce",AZUL_OSC,BLANCO),("Confianza",AZUL_OSC,BLANCO),("% Similitud",AZUL_OSC,BLANCO),
        ("Palabras Comunes",AZUL_CLR,NEGRO),("Total Palabras Unicas",AZUL_CLR,NEGRO),
        ("Material",ROJO_HDR,BLANCO),("Descripcion",ROJO_HDR,BLANCO),("Categoria",ROJO_HDR,BLANCO),
        ("Diferencia",ROJO_HDR,BLANCO),("Costo Unitario",ROJO_HDR,BLANCO),("Costo Total",ROJO_HDR,BLANCO),
        ("Cantidad Cruce",AZUL_OSC,BLANCO),
        ("Material",VERDE_HDR,BLANCO),("Descripcion",VERDE_HDR,BLANCO),("Categoria",VERDE_HDR,BLANCO),
        ("Diferencia",VERDE_HDR,BLANCO),("Costo Unitario",VERDE_HDR,BLANCO),("Costo Total",VERDE_HDR,BLANCO),
        ("Dif. Cantidad",AZUL_MED,BLANCO),("Dif. Costo",AZUL_MED,BLANCO),("Dif. Costo %",AZUL_MED,BLANCO),
    ]
    for i, (h,bg,fg) in enumerate(hdrs_sc, 1):
        write_header(ws_sc, 1, i, h, bg, fg=fg)
    for i, w in enumerate([10,14,12,28,32,18,38,16,14,14,14,12,18,38,16,14,14,14,14,14,12], 1):
        set_col_width(ws_sc, i, w)
    ws_sc.freeze_panes = "A2"

    if not df_cruces_sc.empty:
        for r_idx, row in df_cruces_sc.iterrows():
            er  = r_idx + 2; alt = r_idx % 2 == 0
            sim_val = row["Similitud"]
            _, _, color_sem = nivel_confianza(sim_val / 100, ALTA_MIN, MEDIA_MIN)
            bg_f = ROJO_ALT if alt else ROJO_CLR
            bg_s = VERDE_ALT if alt else VERDE_CLR
            dif_costo = row["Dif_Costo"]; bg_dc = VERDE_CLR if abs(dif_costo) < 0.01 else AMAR_CLR
            dif_pct   = row["Dif_Costo_Pct"]
            bg_pct = ROJO_CLR if dif_pct > 20 else AMAR_CLR if dif_pct > 10 else VERDE_CLR
            vals_sc = [
                (row["No_Cruce"],AMAR_CLR,"center",None),(row["Confianza"],color_sem,"center",None),
                (str(sim_val)+"%",color_sem,"center",None),(row["Comunes"],AZUL_CLR,"left",None),
                (row["Unicas"],AZUL_CLR,"left",None),(row["Mat_F"],bg_f,"left",None),
                (row["Desc_F"],bg_f,"left",None),(row["Cat_F"],bg_f,"left",None),
                (row["Dif_F"],bg_f,"center","#,##0.00"),(row["CostoU_F"],bg_f,"center","#,##0.00"),
                (row["CostoT_F"],bg_f,"center","#,##0.00"),(row["Cantidad_Cruce"],AZUL_OSC,"center","#,##0.00"),
                (row["Mat_S"],bg_s,"left",None),(row["Desc_S"],bg_s,"left",None),
                (row["Cat_S"],bg_s,"left",None),(row["Dif_S"],bg_s,"center","#,##0.00"),
                (row["CostoU_S"],bg_s,"center","#,##0.00"),(row["CostoT_S"],bg_s,"center","#,##0.00"),
                (row["Dif_Cantidad"],VERDE_CLR,"center","#,##0.00"),(dif_costo,bg_dc,"center","#,##0.00"),
                (str(dif_pct)+"%",bg_pct,"center",None),
            ]
            for c_idx, (val,bg,ha,nf) in enumerate(vals_sc, 1):
                c = write_cell(ws_sc, er, c_idx, val, bg=bg, h=ha)
                if nf: c.number_format = nf

    # ── Hoja 4: Sin_Cruce ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Sin_Cruce")
    ws3.sheet_view.showGridLines = False
    ws3.row_dimensions[1].height = 30
    ws3.row_dimensions[2].height = 20
    ws3.merge_cells("A1:H1")
    write_header(ws3, 1, 1, "ITEMS SIN CRUCE POSIBLE", AZUL_OSC, sz=12)
    hdrs3 = ["Material","Descripcion","Categoria","Diferencia","Costo Unitario","Costo Total","Tipo","Cantidad Disponible"]
    for i, h in enumerate(hdrs3, 1):
        write_header(ws3, 2, i, h, AZUL_MED)
    for i, w in enumerate([18,45,18,14,16,16,12,20], 1):
        set_col_width(ws3, i, w)
    if not df_no_cruce.empty:
        for r_idx, row in df_no_cruce.iterrows():
            er = r_idx + 3
            bg = ROJO_CLR if row["Tipo"] == "Faltante" else VERDE_CLR
            vals = [row["Material"],row["Descripcion"],row["Categoria"],
                    row["Diferencia"],row["Costo_Unitario"],row["Costo_Total"],
                    row["Tipo"],row["Cantidad_Disponible"]]
            for c_idx, val in enumerate(vals, 1):
                c = write_cell(ws3, er, c_idx, val, bg=bg)
                if c_idx in (4,5,6,8): c.number_format = '#,##0.00'
    ws3.freeze_panes = "A3"

    # ── Hoja Resumen ──────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Resumen")
    ws4.sheet_view.showGridLines = False
    set_col_width(ws4, 1, 38); set_col_width(ws4, 2, 24)

    def res_titulo(row, texto):
        ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws4.cell(row=row, column=1, value=texto)
        c.font = font(bold=True, color=BLANCO, sz=11)
        c.fill = fill(AZUL_OSC); c.alignment = alig("center"); c.border = brd
        ws4.row_dimensions[row].height = 22

    def res_fila(row, label, value, bg_label=GRIS_CLR, bg_val=BLANCO, num_fmt=None, bold_val=False):
        c1 = ws4.cell(row=row, column=1, value=label)
        c1.font = font(bold=True, sz=10); c1.fill = fill(bg_label)
        c1.border = brd; c1.alignment = alig("left")
        c2 = ws4.cell(row=row, column=2, value=value)
        c2.font = font(bold=bold_val, sz=10); c2.fill = fill(bg_val)
        c2.border = brd; c2.alignment = alig("right")
        if num_fmt: c2.number_format = num_fmt
        ws4.row_dimensions[row].height = 18

    now_str       = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    total_cruces  = len(df_cruces)
    total_sc      = len(df_cruces_sc)
    alta  = len(df_cruces[df_cruces["Confianza"].str.contains("ALTA")])  if not df_cruces.empty else 0
    media = len(df_cruces[df_cruces["Confianza"].str.contains("MEDIA")]) if not df_cruces.empty else 0
    baja  = len(df_cruces[df_cruces["Confianza"].str.contains("BAJA")])  if not df_cruces.empty else 0
    val_cruzado_f = df_cruces["CostoT_F"].sum() if not df_cruces.empty else 0
    val_cruzado_s = df_cruces["CostoT_S"].sum() if not df_cruces.empty else 0
    df_nc_f   = df_no_cruce[df_no_cruce["Tipo"] == "Faltante"] if not df_no_cruce.empty else pd.DataFrame()
    df_nc_s   = df_no_cruce[df_no_cruce["Tipo"] == "Sobrante"] if not df_no_cruce.empty else pd.DataFrame()
    val_sin_f  = df_nc_f["Costo_Total"].sum() if not df_nc_f.empty else 0
    val_sin_s  = df_nc_s["Costo_Total"].sum() if not df_nc_s.empty else 0
    items_sin_f = len(df_nc_f); items_sin_s = len(df_nc_s)
    mins = int(tiempo_seg // 60); segs = round(tiempo_seg % 60, 1)
    tiempo_str = (str(mins) + " min " if mins > 0 else "") + str(segs) + " seg"

    r = 1
    res_titulo(r, "RESUMEN DE CRUCES DE INVENTARIO - GOH"); r += 1
    if info.get("nombre_tienda"):    res_fila(r, "Nombre Tienda",    info["nombre_tienda"],    bg_val=AZUL_CLR); r += 1
    if info.get("centro_sap"):       res_fila(r, "Centro SAP",       info["centro_sap"],        bg_val=AZUL_CLR); r += 1
    if info.get("fecha_inventario"): res_fila(r, "Fecha Inventario", info["fecha_inventario"],  bg_val=AZUL_CLR); r += 1
    res_fila(r, "Fecha de generacion",     now_str,    bg_val=AZUL_CLR); r += 1
    res_fila(r, "Tiempo de procesamiento", tiempo_str, bg_val=AZUL_CLR); r += 1

    r += 1; res_titulo(r, "PARAMETROS UTILIZADOS"); r += 1
    res_fila(r, "Tolerancia Costo %",      f"{int(cfg['TOLERANCIA_COSTO']*100)}%"); r += 1
    res_fila(r, "Umbral Similitud %",       f"{int(cfg['UMBRAL_JACCARD']*100)}%"); r += 1
    res_fila(r, "Umbral Confianza Alta %",  f"{int(ALTA_MIN*100)}%"); r += 1
    res_fila(r, "Umbral Confianza Media %", f"{int(MEDIA_MIN*100)}%"); r += 1
    res_fila(r, "Filtrar por Categoria",    "SI" if cfg["FILTRAR_CATEGORIA"] else "NO"); r += 1

    r += 1; res_titulo(r, "CRUCES GENERADOS"); r += 1
    res_fila(r, "Total cruces (con costo)", total_cruces, bg_val=AMAR_CLR, bold_val=True); r += 1
    res_fila(r, "Total cruces (sin costo)", total_sc,     bg_val=AMAR_CLR); r += 1
    res_fila(r, "Confianza ALTA  🟢",  alta,  bg_val=VERDE_CLR); r += 1
    res_fila(r, "Confianza MEDIA 🟡",  media, bg_val=AMAR_CLR);  r += 1
    res_fila(r, "Confianza BAJA  🔴",  baja,  bg_val=ROJO_CLR);  r += 1

    r += 1; res_titulo(r, "VALOR CRUZADO"); r += 1
    res_fila(r, "Valor faltante cruzado", val_cruzado_f, bg_val=ROJO_CLR,  num_fmt="#,##0.00", bold_val=True); r += 1
    res_fila(r, "Valor sobrante cruzado", val_cruzado_s, bg_val=VERDE_CLR, num_fmt="#,##0.00", bold_val=True); r += 1
    res_fila(r, "Diferencia neta", round(val_cruzado_f + val_cruzado_s, 2), bg_val=AZUL_CLR, num_fmt="#,##0.00", bold_val=True); r += 1

    r += 1; res_titulo(r, "ITEMS SIN CRUCE"); r += 1
    res_fila(r, "Items faltantes sin cruce", items_sin_f, bg_val=ROJO_CLR);  r += 1
    res_fila(r, "Valor faltante sin cruce",  val_sin_f,   bg_val=ROJO_CLR,  num_fmt="#,##0.00"); r += 1
    res_fila(r, "Items sobrantes sin cruce", items_sin_s, bg_val=VERDE_CLR); r += 1
    res_fila(r, "Valor sobrante sin cruce",  val_sin_s,   bg_val=VERDE_CLR, num_fmt="#,##0.00"); r += 1

    wb.move_sheet("Resumen", offset=-wb.index(wb["Resumen"]))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, {
        "total_cruces": total_cruces, "total_sc": total_sc,
        "alta": alta, "media": media, "baja": baja,
        "val_cruzado_f": val_cruzado_f, "val_cruzado_s": val_cruzado_s,
        "items_sin_f": items_sin_f, "items_sin_s": items_sin_s,
        "val_sin_f": val_sin_f, "val_sin_s": val_sin_s,
    }

# ══════════════════════════════════════════════════════════════════════════════
# GENERACION PLANTILLA
# ══════════════════════════════════════════════════════════════════════════════
def generar_plantilla():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Diferencias"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    ws.merge_cells("A1:F1")
    write_header(ws, 1, 1, "CONTRALORIA GOH - DIFERENCIAS DE INVENTARIO", AZUL_OSC, sz=12)
    hdrs = ["Material","Descripción","Categoría","Diferencia","Costo Unitario","Costo Total"]
    for i, h in enumerate(hdrs, 1):
        write_header(ws, 2, i, h, AZUL_MED)
    for i, w in enumerate([18, 45, 18, 14, 16, 16], 1):
        set_col_width(ws, i, w)
    ws.freeze_panes = "A3"

    wc = wb.create_sheet("Configuracion")
    wc.sheet_view.showGridLines = False
    set_col_width(wc, 1, 30); set_col_width(wc, 2, 12)
    set_col_width(wc, 3, 45); set_col_width(wc, 4, 55)
    wc.row_dimensions[1].height = 25
    wc.merge_cells("A1:D1")
    c = wc.cell(row=1, column=1, value="DATOS DEL INVENTARIO")
    c.font = font(bold=True, color=BLANCO, sz=11)
    c.fill = fill(AZUL_OSC); c.alignment = alig("center"); c.border = brd

    datos_tienda = [
        ("Nombre Tienda", "", "Nombre de la tienda auditada"),
        ("Centro SAP",    "", "Codigo del centro en SAP"),
        ("Fecha Inventario", "", "Fecha de toma de inventario (DD/MM/AAAA)"),
    ]
    for i, (param, val, desc) in enumerate(datos_tienda, start=2):
        wc.row_dimensions[i].height = 18
        c1 = wc.cell(row=i, column=1, value=param)
        c1.font = font(bold=True, sz=10); c1.fill = fill(GRIS_CLR); c1.border = brd; c1.alignment = alig("left")
        c2 = wc.cell(row=i, column=2, value=val)
        c2.font = font(sz=10); c2.fill = fill(AMAR_CLR); c2.border = brd; c2.alignment = alig("left")
        c3 = wc.cell(row=i, column=3, value=desc)
        c3.font = font(sz=9, color="595959"); c3.fill = fill(BLANCO); c3.border = brd; c3.alignment = alig("left")
        wc.merge_cells(start_row=i, start_column=3, end_row=i, end_column=4)

    fila_sep = len(datos_tienda) + 2
    wc.row_dimensions[fila_sep].height = 8
    fila_hdr = fila_sep + 1
    wc.row_dimensions[fila_hdr].height = 25
    wc.merge_cells(f"A{fila_hdr}:D{fila_hdr}")
    c = wc.cell(row=fila_hdr, column=1, value="PARAMETROS DE CALCULO")
    c.font = font(bold=True, color=BLANCO, sz=11)
    c.fill = fill(AZUL_OSC); c.alignment = alig("center"); c.border = brd

    fila_sub = fila_hdr + 1
    wc.row_dimensions[fila_sub].height = 20
    for i, h in enumerate(["Parámetro","Valor","Descripción","Interpretación"], 1):
        c = wc.cell(row=fila_sub, column=i, value=h)
        c.font = font(bold=True, color=BLANCO, sz=10)
        c.fill = fill(AZUL_MED); c.border = brd; c.alignment = alig("center")

    parametros = [
        ("Tolerancia Costo %",       5,  "Variación máxima permitida entre costos unitarios",
         "Un ítem de $100 acepta sobrantes entre $95 y $105"),
        ("Umbral Similitud %",       30, "Coincidencia mínima de palabras entre descripciones",
         "De cada 10 palabras únicas entre ambas descripciones, al menos 3 deben coincidir"),
        ("Umbral Confianza Alta %",  70, "Similitud desde la cual el cruce es 🟢 ALTA",
         "7 de cada 10 palabras coinciden — descripciones muy parecidas, cruce confiable"),
        ("Umbral Confianza Media %", 45, "Similitud desde la cual el cruce es 🟡 MEDIA. Por debajo: 🔴 BAJA",
         "Entre 4 y 7 palabras coinciden — revisar antes de aprobar. Menos de 4: revisión obligatoria"),
        ("Filtrar por Categoria",   "SI","Faltante y sobrante deben coincidir en categoría cuando ambos la tienen",
         "SI = BICICLETA solo cruza con BICICLETA. NO = ignora la categoría completamente"),
        ("Generar Diagnostico",     "NO","Genera archivo diagnostico con el detalle de cada decisión",
         "SI = más lento pero permite investigar por qué un ítem no cruzó"),
    ]
    fila_dat = fila_sub + 1
    for i, (param, val, desc, interp) in enumerate(parametros):
        r = fila_dat + i
        wc.row_dimensions[r].height = 30
        alt = i % 2 == 0; bg_row = GRIS_CLR if alt else BLANCO
        c1 = wc.cell(row=r, column=1, value=param)
        c1.font = font(bold=True, sz=10); c1.fill = fill(bg_row); c1.border = brd; c1.alignment = alig("left")
        c2 = wc.cell(row=r, column=2, value=val)
        c2.font = font(bold=True, sz=11); c2.fill = fill(AMAR_CLR); c2.border = brd; c2.alignment = alig("center")
        c3 = wc.cell(row=r, column=3, value=desc)
        c3.font = font(sz=9); c3.fill = fill(bg_row); c3.border = brd; c3.alignment = alig("left")
        c4 = wc.cell(row=r, column=4, value=interp)
        c4.font = font(sz=9, color="375623"); c4.fill = fill(bg_row); c4.border = brd; c4.alignment = alig("left")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ══════════════════════════════════════════════════════════════════════════════
# UI PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
st.title("📦 Motor de Cruces de Inventario")
st.caption("Contraloría GOH — Procesamiento automático de diferencias de inventario")

# Sidebar — descarga plantilla
with st.sidebar:
    st.header("⚙️ Instrucciones")
    st.markdown("""
    **Pasos:**
    1. Descarga la plantilla
    2. Completa la hoja **Diferencias** con tus datos
    3. Ajusta los parámetros en la hoja **Configuracion**
    4. Sube el archivo y ejecuta el proceso
    """)
    st.divider()
    st.subheader("📥 Plantilla")
    plantilla_bytes = generar_plantilla()
    st.download_button(
        label="⬇️ Descargar Plantilla",
        data=plantilla_bytes,
        file_name="Diferencias_Plantilla.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.caption("Llena la plantilla y súbela abajo para procesar.")

# Área principal
uploaded_file = st.file_uploader(
    "📂 Sube tu archivo Diferencias.xlsx",
    type=["xlsx"],
    help="El archivo debe contener las hojas 'Diferencias' y 'Configuracion'"
)

if uploaded_file:
    file_bytes = uploaded_file.read()

    # Vista previa de configuración
    try:
        cfg = cargar_configuracion(io.BytesIO(file_bytes))
        df_prev = cargar_datos(io.BytesIO(file_bytes))

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total registros", len(df_prev))
        with col2:
            st.metric("🔴 Faltantes", len(df_prev[df_prev["Diferencia"] < 0]))
        with col3:
            st.metric("🟢 Sobrantes", len(df_prev[df_prev["Diferencia"] > 0]))

        if cfg["nombre_tienda"] or cfg["centro_sap"]:
            st.info(f"📍 Tienda: **{cfg['nombre_tienda']}**  |  Centro SAP: **{cfg['centro_sap']}**  |  Fecha: **{cfg['fecha_inventario']}**")

        with st.expander("🔧 Parámetros cargados desde el archivo"):
            pc1, pc2 = st.columns(2)
            with pc1:
                st.write(f"**Tolerancia costo:** {int(cfg['TOLERANCIA_COSTO']*100)}%")
                st.write(f"**Umbral similitud:** {int(cfg['UMBRAL_JACCARD']*100)}%")
                st.write(f"**Confianza ALTA desde:** {int(cfg['ALTA_MIN']*100)}%")
            with pc2:
                st.write(f"**Confianza MEDIA desde:** {int(cfg['MEDIA_MIN']*100)}%")
                st.write(f"**Filtrar por categoría:** {'Sí' if cfg['FILTRAR_CATEGORIA'] else 'No'}")

        st.divider()

        if st.button("🚀 Ejecutar Cruces de Inventario", type="primary", use_container_width=True):
            progress_bar = st.progress(0, text="Iniciando proceso...")
            status_text  = st.empty()
            ts_inicio    = time.time()

            try:
                df_c, df_csc, df_nc, rf, rs = ejecutar_cruces(
                    df_prev.copy(), cfg, progress_bar, status_text
                )
                tiempo_total = time.time() - ts_inicio
                progress_bar.progress(1.0, text="✅ Proceso completado")
                status_text.empty()

                excel_output, stats = generar_excel(
                    df_prev, df_c, df_csc, df_nc, rf, rs, tiempo_total, cfg
                )

                # Resumen de resultados
                st.success(f"✅ Proceso completado en {round(tiempo_total, 1)} segundos")

                st.subheader("📊 Resumen de Resultados")
                r1, r2, r3 = st.columns(3)
                with r1:
                    st.metric("Cruces con costo", stats["total_cruces"])
                    st.metric("Cruces sin costo", stats["total_sc"])
                with r2:
                    st.metric("🟢 Confianza ALTA",  stats["alta"])
                    st.metric("🟡 Confianza MEDIA", stats["media"])
                    st.metric("🔴 Confianza BAJA",  stats["baja"])
                with r3:
                    st.metric("Faltantes sin cruce", stats["items_sin_f"])
                    st.metric("Sobrantes sin cruce", stats["items_sin_s"])

                st.divider()
                timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                st.download_button(
                    label="⬇️ Descargar Resultado Excel",
                    data=excel_output,
                    file_name=f"Cruces_Inventario_GOH_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                )

            except Exception as e:
                progress_bar.empty()
                st.error(f"❌ Error durante el procesamiento: {str(e)}")

    except ValueError as e:
        st.error(f"❌ Error al leer el archivo: {str(e)}")
    except Exception as e:
        st.error(f"❌ Error inesperado: {str(e)}")

else:
    st.info("👆 Sube tu archivo Excel para comenzar. Si aún no tienes la plantilla, descárgala desde el panel izquierdo.")
